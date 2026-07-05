#!/usr/bin/env python3
"""
cli_usage_tracker.py — Track CLI vs NETCONF commit usage across network devices.

Queries commit history from Juniper, Nokia, and Arrcus devices to analyze
whether commits were performed via CLI or NETCONF. Generates detailed reports
showing usage patterns per device, per user, and fleet-wide statistics.

Vendor Support:
  - Juniper: NETCONF over SSH (port 22) - RPC <get-commit-information> (fallback to CLI)
  - Nokia:   NETCONF over SSH (port 22) - YANG commit-history (fallback to CLI)
  - Arrcus:  NETCONF over SSH (port 22) - commit-list query (fallback to CLI)

Usage:
    # Pull device list dynamically from ISIS (recommended - no inventory file needed)
    python3 cli_usage_tracker.py --from-isis

    # Use specific ISIS router for discovery
    python3 cli_usage_tracker.py --from-isis --isis-router router1.lab

    # Use custom .env file for credentials (avoids prompts)
    python3 cli_usage_tracker.py --from-isis --env-file ~/my-credentials.env

    # Auto-detect inventory (CSV from build_inventory.py or existing Excel)
    python3 cli_usage_tracker.py

    # Use specific inventory file
    python3 cli_usage_tracker.py --inventory ~/device_versions.csv

    # Query specific devices or vendors
    python3 cli_usage_tracker.py --from-isis --vendor Juniper --max-commits 50
    python3 cli_usage_tracker.py --devices router1,router2,switch1
    python3 cli_usage_tracker.py --output-dir ./reports

Credentials:
    The script will attempt to load credentials in this order:
    1. From .env file (use --env-file to specify, or auto-detects ./.env or ~/.env)
    2. From environment variables: DEVICE_USERNAME, DEVICE_PASSWORD
    3. Interactive prompts (if not found in .env or environment)

    .env file format (to avoid entering credentials every time):
        DEVICE_USERNAME=your_username
        DEVICE_PASSWORD=your_password

Output:
    Text Reports:
        cli_usage_report_YYYYMMDD_HHMMSS.txt - Detailed per-device report (all commits)
        cli_usage_summary_YYYYMMDD_HHMMSS.txt - Fleet-wide summary statistics
        cli_usage_cli_only_YYYYMMDD_HHMMSS.txt - CLI commits only (NETCONF filtered out)
        cli_usage_report_YYYYMMDD_HHMMSS.csv - CSV export for analysis
    HTML Reports:
        cli_usage_report_YYYYMMDD_HHMMSS.html - Detailed per-device report (all commits)
        cli_usage_summary_YYYYMMDD_HHMMSS.html - Fleet-wide summary with styled tables
        cli_usage_cli_only_YYYYMMDD_HHMMSS.html - CLI commits only (NETCONF filtered out)
"""

from __future__ import annotations

import os
import re
import sys
import time
import csv
import argparse
import getpass
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dotenv import load_dotenv

import logging
import threading

# Suppress paramiko logging noise
logging.getLogger("paramiko").setLevel(logging.CRITICAL)
logging.getLogger("paramiko.transport").setLevel(logging.CRITICAL)

_original_excepthook = threading.excepthook if hasattr(threading, "excepthook") else None

def _suppress_paramiko_threads(args):
    if args.thread and "paramiko" in (args.thread.name or "").lower():
        return
    if _original_excepthook:
        _original_excepthook(args)

if hasattr(threading, "excepthook"):
    threading.excepthook = _suppress_paramiko_threads

# Global variables (will be set in main() with prompts if needed)
USERNAME = None
PASSWORD = None
PORT_SSH = 22

INVENTORY_FILE = os.path.expanduser("~/device_inventory.xlsx")
INVENTORY_CSV = os.path.expanduser("~/device_inventory.csv")

# ISIS discovery defaults
ISIS_ROUTER = "isis-router.lab"  # Change this to your ISIS router hostname
ISIS_PORT = 22

# ---------------------------------------------------------------------------
# ISIS Discovery (Dynamic Device List)
# ---------------------------------------------------------------------------

def pull_isis_hostnames(router: str = ISIS_ROUTER, port: int = ISIS_PORT) -> list[str]:
    """
    Fetch all ISIS L2 hostnames via NETCONF over SSH from a Juniper router.
    Returns a sorted list of unique hostnames.
    """
    from lxml import etree
    from ncclient import manager

    RPC = "<get-isis-database-information><detail/></get-isis-database-information>"

    print(f"Connecting to {router}:{port} via NETCONF over SSH to pull ISIS database...")
    with manager.connect(
        host=router,
        port=port,
        username=USERNAME,
        password=PASSWORD,
        hostkey_verify=False,
        device_params={"name": "junos"},
        timeout=120,
    ) as m:
        print("  Fetching ISIS database (detail)...")
        resp = m.dispatch(etree.fromstring(RPC))

    if hasattr(resp, "data_ele"):
        root = resp.data_ele
    else:
        xml_str = resp.tostring if hasattr(resp, "tostring") else str(resp)
        root = etree.fromstring(xml_str.encode() if isinstance(xml_str, str) else xml_str)

    hostnames = set()
    for elem in root.iter():
        tag = getattr(elem.tag, "localname", None) or (
            re.sub(r"\{.*\}", "", elem.tag) if isinstance(elem.tag, str) else ""
        )
        if tag in ("isis-hostname", "hostname"):
            h = (elem.text or "").strip()
            if h:
                hostnames.add(h)

    # Fallback: parse <lsp-id>
    if not hostnames:
        for elem in root.iter():
            tag = re.sub(r"\{.*\}", "", elem.tag) if isinstance(elem.tag, str) else ""
            if tag == "lsp-id":
                lsp = (elem.text or "").strip()
                m = re.match(r"^(.+?)\.00-\d+$", lsp)
                if m:
                    hostnames.add(m.group(1))

    # Strip _RE0/_RE1 suffixes
    cleaned = set()
    for h in hostnames:
        cleaned.add(re.sub(r"[-_][Rr][Ee]\d+$", "", h))

    result = sorted(cleaned)
    print(f"  Found {len(result)} unique ISIS hostnames")
    return result


# Role patterns for vendor detection (borrowed from build_inventory.py)
ROLE_PATTERNS = [
    (r"^bb\d+\.", "bb", "Juniper"),
    (r"^se\d+\.", "se", "Juniper"),
    (r"^ec-gw\d+\.", "ec_gw", "Juniper"),
    (r"^ecgw\d+\.", "ec_gw", "Juniper"),
    (r"^grr\d+\.", "grr", "Juniper"),
    (r"^mrr\d+\.", "mrr", "Juniper"),
    (r"^ec-mrr\d+\.", "ec_mrr", "Juniper"),
    (r"^ecmrr\d+\.", "ec_mrr", "Juniper"),
    (r"^cor\d+\.", "sp", "Nokia"),
    (r"^sp\d+\.", "sp", "Nokia"),
    (r"^dis\d+\.", "use", "Nokia"),
    (r"^use3\.", "use", "Nokia"),  # use3.* are Nokia SR-OS devices
    (r"^use\d+\.", "use", "Arrcus"),  # Other use*.* are Arrcus
    (r"^nat\d+\.", "nat", "Juniper"),
    (r"^pol\d+\.", "pol", "Juniper"),
    (r"^android[_-]?gw", "android_gw", "Juniper"),
]


def guess_vendor_from_hostname(hostname: str) -> str:
    """
    Guess vendor from hostname prefix based on naming conventions.
    Returns: "Juniper", "Nokia", "Arrcus", or "unknown"
    """
    hn = hostname.lower()
    for pattern, role, vendor in ROLE_PATTERNS:
        if re.match(pattern, hn):
            return vendor
    return "unknown"


# ---------------------------------------------------------------------------
# Inventory loading
# ---------------------------------------------------------------------------

def load_inventory(file_path: str) -> dict[str, str]:
    """
    Load hostname -> vendor mapping from inventory file.

    Supports:
    - CSV files (from build_inventory.py): hostname,vendor,platform,version,role,nos,error
    - Excel files (ISIS_Device_List sheet): hostname,vendor columns
    """
    file_path_obj = Path(file_path)

    # CSV format (from build_inventory.py)
    if file_path_obj.suffix.lower() == ".csv":
        mapping = {}
        with open(file_path, "r") as f:
            reader = csv.DictReader(f)
            for row in reader:
                hostname = row.get("hostname", "").strip()
                vendor = row.get("vendor", "").strip()
                if hostname and vendor:
                    # Normalize vendor names (CSV uses uppercase: JUNIPER, NOKIA, ARRCUS)
                    vendor_normalized = vendor.capitalize() if vendor.upper() in ["JUNIPER", "NOKIA", "ARRCUS"] else vendor
                    mapping[hostname] = vendor_normalized
        return mapping

    # Excel format (ISIS_Device_List sheet)
    try:
        from openpyxl import load_workbook
    except ImportError:
        import pandas as pd
        df = pd.read_excel(file_path, sheet_name="ISIS_Device_List")
        return dict(zip(df["hostname"], df["vendor"]))

    wb = load_workbook(file_path, read_only=True, data_only=True)
    if "ISIS_Device_List" in wb.sheetnames:
        ws = wb["ISIS_Device_List"]
    else:
        ws = wb.active

    rows = ws.iter_rows(values_only=True)
    header = next(rows)

    # Find column indices (case-insensitive)
    h_idx = None
    v_idx = None
    for i, col in enumerate(header):
        if col and str(col).strip().lower() == "hostname":
            h_idx = i
        if col and str(col).strip().lower() == "vendor":
            v_idx = i

    if h_idx is None or v_idx is None:
        print(f"ERROR: Could not find hostname/vendor columns. Found: {header}", file=sys.stderr)
        sys.exit(1)

    mapping = {}
    for row in rows:
        hostname = row[h_idx]
        vendor = row[v_idx]
        if hostname and vendor:
            mapping[str(hostname).strip()] = str(vendor).strip()

    wb.close()
    return mapping


# ---------------------------------------------------------------------------
# Vendor-specific commit queries (returning multiple commits)
# ---------------------------------------------------------------------------

def query_juniper_netconf(hostname: str, max_commits: int = 30) -> list[dict]:
    """Query Juniper via NETCONF for commit history."""
    from ncclient import manager
    from lxml import etree

    RPC = "<get-commit-information></get-commit-information>"

    with manager.connect(
        host=hostname,
        port=PORT_SSH,
        username=USERNAME,
        password=PASSWORD,
        hostkey_verify=False,
        device_params={"name": "junos"},
        timeout=30,
    ) as m:
        resp = m.dispatch(etree.fromstring(RPC))

    if hasattr(resp, "data_ele"):
        root = resp.data_ele
    else:
        xml_str = resp.tostring if hasattr(resp, "tostring") else str(resp)
        root = etree.fromstring(
            xml_str.encode() if isinstance(xml_str, str) else xml_str
        )

    # Parse all commit-history entries
    commits = []
    for elem in root.iter():
        tag = etree.QName(elem.tag).localname if isinstance(elem.tag, str) else ""
        if tag == "commit-history":
            dt = user = via = log = ""
            for child in elem:
                ctag = etree.QName(child.tag).localname if isinstance(child.tag, str) else ""
                if ctag == "date-time":
                    dt = (child.text or "").strip()
                elif ctag == "user":
                    user = (child.text or "").strip()
                elif ctag == "client":
                    via = (child.text or "").strip()
                elif ctag == "log":
                    log = (child.text or "").strip()

            if dt and user and via:
                commits.append({
                    "datetime": dt,
                    "user": user,
                    "via": via.lower(),  # Normalize to lowercase
                    "comment": log,
                })

            if len(commits) >= max_commits:
                break

    return commits


def query_juniper_ssh(hostname: str, max_commits: int = 30) -> list[dict]:
    """Fallback: Query Juniper via SSH for commit history."""
    import paramiko

    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    client.connect(
        hostname, port=PORT_SSH, username=USERNAME, password=PASSWORD,
        look_for_keys=False, allow_agent=False, timeout=30,
    )
    stdin, stdout, stderr = client.exec_command("show system commit", timeout=30)
    output = stdout.read().decode("utf-8", errors="replace")
    client.close()

    # Parse: "0   2026-06-18 18:32:25 UTC by vimanalo via netconf commit synchronize"
    commits = []
    for line in output.strip().splitlines():
        m = re.match(
            r'\s*\d+\s+(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}\s+\S+)\s+by\s+(\S+)\s+via\s+(\S+)',
            line
        )
        if m:
            commits.append({
                "datetime": m.group(1),
                "user": m.group(2),
                "via": m.group(3).lower(),  # Normalize to lowercase
                "comment": line.split("via " + m.group(3))[-1].strip() if "via " + m.group(3) in line else "",
            })

            if len(commits) >= max_commits:
                break

    return commits


def query_juniper(hostname: str, max_commits: int = 30) -> list[dict]:
    """Query Juniper — try NETCONF first, fall back to SSH."""
    try:
        return query_juniper_netconf(hostname, max_commits)
    except Exception:
        return query_juniper_ssh(hostname, max_commits)


def query_nokia_netconf(hostname: str, max_commits: int = 30) -> list[dict]:
    """Query Nokia SR-OS via NETCONF for commit history."""
    from ncclient import manager
    from lxml import etree

    # Try multiple NETCONF approaches for Nokia
    # Approach 1: state/system/management-interface
    RPC = """
    <get>
        <filter type="subtree">
            <state xmlns="urn:nokia.com:sros:ns:yang:sr:state">
                <system>
                    <management-interface>
                        <commit-history/>
                    </management-interface>
                </system>
            </state>
        </filter>
    </get>
    """

    with manager.connect(
        host=hostname,
        port=PORT_SSH,
        username=USERNAME,
        password=PASSWORD,
        hostkey_verify=False,
        device_params={"name": "sros"},
        timeout=30,
    ) as m:
        resp = m.dispatch(etree.fromstring(RPC))

    if hasattr(resp, "data_ele"):
        root = resp.data_ele
    else:
        xml_str = resp.tostring if hasattr(resp, "tostring") else str(resp)
        root = etree.fromstring(
            xml_str.encode() if isinstance(xml_str, str) else xml_str
        )

    # Parse commit history entries
    commits = []
    for elem in root.iter():
        tag = etree.QName(elem.tag).localname if isinstance(elem.tag, str) else ""
        if tag in ("commit-history", "commit"):
            commit_id = timestamp = user = method = ""
            for child in elem:
                ctag = etree.QName(child.tag).localname if isinstance(child.tag, str) else ""
                if ctag in ("commit-id", "id"):
                    commit_id = (child.text or "").strip()
                elif ctag == "timestamp":
                    timestamp = (child.text or "").strip()
                elif ctag == "user":
                    user = (child.text or "").strip()
                elif ctag in ("method", "client"):
                    method = (child.text or "").strip()

            if timestamp and user and method:
                commits.append({
                    "datetime": timestamp,
                    "user": user,
                    "via": method.lower(),
                    "comment": "",
                })

            if len(commits) >= max_commits:
                break

    # If NETCONF didn't return commits, raise exception to trigger SSH fallback
    if not commits:
        raise Exception("NETCONF returned no commits")

    return commits


def query_nokia_ssh(hostname: str, max_commits: int = 30) -> list[dict]:
    """Fallback: Query Nokia SR-OS via SSH for commit history."""
    import paramiko

    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    client.connect(
        hostname, port=PORT_SSH, username=USERNAME, password=PASSWORD,
        look_for_keys=False, allow_agent=False, timeout=30,
    )

    # Nokia MD-CLI needs an interactive shell
    shell = client.invoke_shell()
    time.sleep(1.5)
    shell.recv(65535)  # clear banner

    shell.send("environment more false\n")
    time.sleep(0.8)
    shell.recv(65535)

    # Try both MD-CLI and classic CLI commands
    commands = [
        "show system management-interface commit-history",  # MD-CLI
        "/show system management-interface commit-history",  # MD-CLI with explicit path
    ]

    text = ""
    for cmd in commands:
        shell.send(f"{cmd}\n")
        time.sleep(3)
        output = b""
        while shell.recv_ready():
            output += shell.recv(65535)
            time.sleep(0.5)

        text = output.decode("utf-8", errors="replace")

        # Check if we got valid output
        if "Committed" in text or "commit" in text.lower():
            break

    client.close()

    # Parse commit entries - try multiple patterns
    commits = []

    # Pattern 1: "203   Committed 2026-05-05T18:21:58.7+00:00 by username (NETCONF) from ..."
    for line in text.splitlines():
        m = re.search(
            r'(\d+)\s+Committed\s+(\S+)\s+by\s+(\S+)\s+\(([^)]+)\)',
            line
        )
        if m:
            commits.append({
                "datetime": m.group(2),
                "user": m.group(3),
                "via": m.group(4).lower(),  # Normalize to lowercase
                "comment": "",
            })

            if len(commits) >= max_commits:
                break

    # Pattern 2: If pattern 1 didn't work, try simpler pattern
    if not commits:
        for line in text.splitlines():
            # Try to match any line with timestamp, user, and method indicators
            if re.search(r'\d{4}-\d{2}-\d{2}', line) and ('netconf' in line.lower() or 'cli' in line.lower()):
                # Extract what we can
                timestamp_match = re.search(r'(\d{4}-\d{2}-\d{2}[T\s]\d{2}:\d{2}:\d{2}[^\s]*)', line)
                if timestamp_match:
                    timestamp = timestamp_match.group(1)
                    # Try to extract user and method
                    parts = line.split()
                    user = "unknown"
                    method = "cli"

                    if 'netconf' in line.lower():
                        method = "netconf"
                    if 'by' in line.lower():
                        idx = line.lower().index('by')
                        after_by = line[idx+2:].strip().split()[0]
                        if after_by:
                            user = after_by

                    commits.append({
                        "datetime": timestamp,
                        "user": user,
                        "via": method,
                        "comment": "",
                    })

                    if len(commits) >= max_commits:
                        break

    return commits


def query_nokia(hostname: str, max_commits: int = 30) -> list[dict]:
    """Query Nokia — try NETCONF first, fall back to SSH."""
    try:
        return query_nokia_netconf(hostname, max_commits)
    except Exception:
        return query_nokia_ssh(hostname, max_commits)


def query_arrcus_netconf(hostname: str, max_commits: int = 30) -> list[dict]:
    """Query Arrcus via NETCONF for commit history."""
    from ncclient import manager
    from lxml import etree

    # Arrcus NETCONF RPC for commit list
    RPC = """
    <get>
        <filter type="subtree">
            <commit-list xmlns="http://www.arrcus.com"/>
        </filter>
    </get>
    """

    with manager.connect(
        host=hostname,
        port=PORT_SSH,
        username=USERNAME,
        password=PASSWORD,
        hostkey_verify=False,
        device_params={"name": "default"},
        timeout=30,
    ) as m:
        resp = m.dispatch(etree.fromstring(RPC))

    if hasattr(resp, "data_ele"):
        root = resp.data_ele
    else:
        xml_str = resp.tostring if hasattr(resp, "tostring") else str(resp)
        root = etree.fromstring(
            xml_str.encode() if isinstance(xml_str, str) else xml_str
        )

    # Parse commit entries
    commits = []
    for elem in root.iter():
        tag = etree.QName(elem.tag).localname if isinstance(elem.tag, str) else ""
        if tag == "commit":
            commit_id = user = client = timestamp = ""
            for child in elem:
                ctag = etree.QName(child.tag).localname if isinstance(child.tag, str) else ""
                if ctag == "id":
                    commit_id = (child.text or "").strip()
                elif ctag == "user":
                    user = (child.text or "").strip()
                elif ctag == "client":
                    client = (child.text or "").strip()
                elif ctag == "timestamp":
                    timestamp = (child.text or "").strip()

            if timestamp and user and client:
                commits.append({
                    "datetime": timestamp,
                    "user": user,
                    "via": client.lower(),
                    "comment": "",
                })

            if len(commits) >= max_commits:
                break

    return commits


def query_arrcus_ssh(hostname: str, max_commits: int = 30) -> list[dict]:
    """Fallback: Query Arrcus via SSH for commit list."""
    import paramiko

    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    client.connect(
        hostname, port=PORT_SSH, username=USERNAME, password=PASSWORD,
        look_for_keys=False, allow_agent=False, timeout=30,
    )
    stdin, stdout, stderr = client.exec_command("show configuration commit list", timeout=30)
    output = stdout.read().decode("utf-8", errors="replace")
    client.close()

    # Parse: "0    10081    svc-aos-pr netconf     2026-05-07 16:59:52"
    # Format: SNo  ID  User  Client  TimeStamp  Label  Comment
    commits = []
    for line in output.strip().splitlines():
        # Skip header lines
        if line.startswith("SNo") or line.startswith("~~~") or not line.strip():
            continue

        m = re.match(
            r'\s*\d+\s+\d+\s+(\S+)\s+(\S+)\s+(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})',
            line
        )
        if m:
            commits.append({
                "datetime": m.group(3),
                "user": m.group(1),
                "via": m.group(2).lower(),  # Normalize to lowercase
                "comment": "",
            })

            if len(commits) >= max_commits:
                break

    return commits


def query_arrcus(hostname: str, max_commits: int = 30) -> list[dict]:
    """Query Arrcus — try NETCONF first, fall back to SSH."""
    try:
        return query_arrcus_netconf(hostname, max_commits)
    except Exception:
        return query_arrcus_ssh(hostname, max_commits)


# ---------------------------------------------------------------------------
# Main query orchestration
# ---------------------------------------------------------------------------

def query_device(hostname: str, vendor: str, max_commits: int = 30, debug: bool = False) -> dict:
    """Query a single device based on its vendor."""
    netconf_error = None
    ssh_error = None

    try:
        if vendor == "Juniper":
            commits = query_juniper(hostname, max_commits)
        elif vendor == "Nokia":
            commits = query_nokia(hostname, max_commits)
        elif vendor == "Arrcus":
            commits = query_arrcus(hostname, max_commits)
        else:
            return {"hostname": hostname, "vendor": vendor, "error": f"Unknown vendor: {vendor}", "commits": []}

        if not commits:
            return {"hostname": hostname, "vendor": vendor, "error": "No commits found (both NETCONF and SSH returned empty)", "commits": []}

        return {"hostname": hostname, "vendor": vendor, "error": None, "commits": commits}

    except Exception as e:
        error_msg = str(e)[:100]
        if debug:
            import traceback
            error_msg = f"{error_msg} | {traceback.format_exc()[:200]}"
        return {"hostname": hostname, "vendor": vendor, "error": error_msg, "commits": []}


# ---------------------------------------------------------------------------
# Report generation
# ---------------------------------------------------------------------------

def analyze_commits(results: list[dict]) -> dict:
    """Analyze commit data and generate statistics."""
    stats = {
        "total_devices": len(results),
        "devices_queried": sum(1 for r in results if not r["error"]),
        "devices_failed": sum(1 for r in results if r["error"]),
        "total_commits": sum(len(r["commits"]) for r in results),
        "cli_commits": 0,
        "netconf_commits": 0,
        "other_commits": 0,
        "by_vendor": defaultdict(lambda: {"cli": 0, "netconf": 0, "other": 0, "total": 0}),
        "by_user": defaultdict(lambda: {"cli": 0, "netconf": 0, "other": 0, "total": 0}),
    }

    for result in results:
        vendor = result["vendor"]
        for commit in result["commits"]:
            via = commit["via"]
            user = commit["user"]

            # Categorize commit method
            if "cli" in via:
                stats["cli_commits"] += 1
                stats["by_vendor"][vendor]["cli"] += 1
                stats["by_user"][user]["cli"] += 1
            elif "netconf" in via or "grpc" in via:
                stats["netconf_commits"] += 1
                stats["by_vendor"][vendor]["netconf"] += 1
                stats["by_user"][user]["netconf"] += 1
            else:
                stats["other_commits"] += 1
                stats["by_vendor"][vendor]["other"] += 1
                stats["by_user"][user]["other"] += 1

            stats["by_vendor"][vendor]["total"] += 1
            stats["by_user"][user]["total"] += 1

    return stats


def write_detailed_report(results: list[dict], filepath: Path):
    """Write detailed per-device report."""
    with open(filepath, "w") as f:
        f.write("=" * 120 + "\n")
        f.write("CLI USAGE TRACKER - DETAILED REPORT\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 120 + "\n\n")

        for result in sorted(results, key=lambda r: r["hostname"]):
            hostname = result["hostname"]
            vendor = result["vendor"]
            error = result["error"]
            commits = result["commits"]

            f.write(f"\n{'=' * 120}\n")
            f.write(f"DEVICE: {hostname} (Vendor: {vendor})\n")
            f.write(f"{'=' * 120}\n")

            if error:
                f.write(f"ERROR: {error}\n")
                continue

            if not commits:
                f.write("No commits found\n")
                continue

            # Calculate device-specific stats
            cli_count = sum(1 for c in commits if "cli" in c["via"])
            netconf_count = sum(1 for c in commits if "netconf" in c["via"] or "grpc" in c["via"])
            other_count = len(commits) - cli_count - netconf_count

            f.write(f"\nTotal Commits: {len(commits)}\n")
            f.write(f"  CLI:     {cli_count:4d} ({cli_count/len(commits)*100:5.1f}%)\n")
            f.write(f"  NETCONF: {netconf_count:4d} ({netconf_count/len(commits)*100:5.1f}%)\n")
            if other_count > 0:
                f.write(f"  Other:   {other_count:4d} ({other_count/len(commits)*100:5.1f}%)\n")

            f.write(f"\n{'TIMESTAMP':<30} {'USER':<20} {'METHOD':<12} COMMENT\n")
            f.write("-" * 120 + "\n")

            for commit in commits:
                f.write(
                    f"{commit['datetime']:<30} {commit['user']:<20} {commit['via']:<12} "
                    f"{commit['comment'][:50]}\n"
                )


def write_summary_report(stats: dict, filepath: Path):
    """Write fleet-wide summary statistics."""
    with open(filepath, "w") as f:
        f.write("=" * 100 + "\n")
        f.write("CLI USAGE TRACKER - SUMMARY REPORT\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 100 + "\n\n")

        # Overall statistics
        f.write("FLEET-WIDE STATISTICS\n")
        f.write("-" * 100 + "\n")
        f.write(f"Total Devices:        {stats['total_devices']}\n")
        f.write(f"Devices Queried:      {stats['devices_queried']}\n")
        f.write(f"Devices Failed:       {stats['devices_failed']}\n")
        f.write(f"Total Commits:        {stats['total_commits']}\n\n")

        if stats["total_commits"] > 0:
            cli_pct = stats["cli_commits"] / stats["total_commits"] * 100
            netconf_pct = stats["netconf_commits"] / stats["total_commits"] * 100
            other_pct = stats["other_commits"] / stats["total_commits"] * 100

            f.write(f"CLI Commits:          {stats['cli_commits']:6d} ({cli_pct:5.1f}%)\n")
            f.write(f"NETCONF Commits:      {stats['netconf_commits']:6d} ({netconf_pct:5.1f}%)\n")
            if stats["other_commits"] > 0:
                f.write(f"Other Commits:        {stats['other_commits']:6d} ({other_pct:5.1f}%)\n")

        # By vendor
        f.write("\n\nBREAKDOWN BY VENDOR\n")
        f.write("-" * 100 + "\n")
        f.write(f"{'Vendor':<15} {'Total':>8} {'CLI':>8} {'CLI %':>8} {'NETCONF':>8} {'NETCONF %':>10} {'Other':>8}\n")
        f.write("-" * 100 + "\n")

        for vendor in sorted(stats["by_vendor"].keys()):
            v_stats = stats["by_vendor"][vendor]
            total = v_stats["total"]
            if total > 0:
                cli_pct = v_stats["cli"] / total * 100
                netconf_pct = v_stats["netconf"] / total * 100
                f.write(
                    f"{vendor:<15} {total:8d} {v_stats['cli']:8d} {cli_pct:7.1f}% "
                    f"{v_stats['netconf']:8d} {netconf_pct:9.1f}% {v_stats['other']:8d}\n"
                )

        # By user
        f.write("\n\nBREAKDOWN BY USER (Top 20)\n")
        f.write("-" * 100 + "\n")
        f.write(f"{'User':<25} {'Total':>8} {'CLI':>8} {'CLI %':>8} {'NETCONF':>8} {'NETCONF %':>10} {'Other':>8}\n")
        f.write("-" * 100 + "\n")

        # Sort by total commits descending
        sorted_users = sorted(
            stats["by_user"].items(),
            key=lambda x: x[1]["total"],
            reverse=True
        )[:20]

        for user, u_stats in sorted_users:
            total = u_stats["total"]
            if total > 0:
                cli_pct = u_stats["cli"] / total * 100
                netconf_pct = u_stats["netconf"] / total * 100
                f.write(
                    f"{user:<25} {total:8d} {u_stats['cli']:8d} {cli_pct:7.1f}% "
                    f"{u_stats['netconf']:8d} {netconf_pct:9.1f}% {u_stats['other']:8d}\n"
                )


def write_csv_export(results: list[dict], filepath: Path):
    """Write CSV export for further analysis."""
    with open(filepath, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Hostname", "Vendor", "Timestamp", "User", "Method", "Comment", "Error"])

        for result in sorted(results, key=lambda r: r["hostname"]):
            hostname = result["hostname"]
            vendor = result["vendor"]
            error = result.get("error", "")

            if error:
                writer.writerow([hostname, vendor, "", "", "", "", error])
            elif result["commits"]:
                for commit in result["commits"]:
                    writer.writerow([
                        hostname,
                        vendor,
                        commit["datetime"],
                        commit["user"],
                        commit["via"],
                        commit["comment"],
                        ""
                    ])
            else:
                writer.writerow([hostname, vendor, "", "", "", "", "No commits found"])


def write_cli_only_report(results: list[dict], filepath: Path):
    """Write detailed report showing ONLY CLI commits (filters out NETCONF-only entries)."""
    with open(filepath, "w") as f:
        f.write("=" * 120 + "\n")
        f.write("CLI USAGE TRACKER - CLI-ONLY REPORT\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 120 + "\n")
        f.write("NOTE: This report shows ONLY commits made via CLI (NETCONF commits filtered out)\n")
        f.write("=" * 120 + "\n\n")

        cli_device_count = 0
        total_cli_commits = 0

        for result in sorted(results, key=lambda r: r["hostname"]):
            hostname = result["hostname"]
            vendor = result["vendor"]
            error = result["error"]
            commits = result["commits"]

            if error or not commits:
                continue

            # Filter to only CLI commits
            cli_commits = [c for c in commits if "cli" in c["via"]]

            if not cli_commits:
                continue

            cli_device_count += 1
            total_cli_commits += len(cli_commits)

            f.write(f"\n{'=' * 120}\n")
            f.write(f"DEVICE: {hostname} (Vendor: {vendor})\n")
            f.write(f"{'=' * 120}\n")
            f.write(f"CLI Commits: {len(cli_commits)}\n")

            f.write(f"\n{'TIMESTAMP':<30} {'USER':<20} {'METHOD':<12} COMMENT\n")
            f.write("-" * 120 + "\n")

            for commit in cli_commits:
                f.write(
                    f"{commit['datetime']:<30} {commit['user']:<20} {commit['via']:<12} "
                    f"{commit['comment'][:50]}\n"
                )

        f.write(f"\n{'=' * 120}\n")
        f.write(f"CLI-ONLY SUMMARY\n")
        f.write(f"{'=' * 120}\n")
        f.write(f"Devices with CLI commits: {cli_device_count}\n")
        f.write(f"Total CLI commits:        {total_cli_commits}\n")


def write_detailed_report_html(results: list[dict], filepath: Path):
    """Write detailed per-device report in HTML format."""
    with open(filepath, "w") as f:
        f.write("""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>CLI Usage Tracker - Detailed Report</title>
    <style>
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }
        h1 {
            color: #2c3e50;
            border-bottom: 3px solid #3498db;
            padding-bottom: 10px;
        }
        h2 {
            color: #34495e;
            background-color: #ecf0f1;
            padding: 10px;
            border-left: 5px solid #3498db;
            margin-top: 30px;
        }
        .meta {
            color: #7f8c8d;
            font-style: italic;
            margin-bottom: 20px;
        }
        .device-section {
            background-color: white;
            border: 1px solid #ddd;
            border-radius: 5px;
            padding: 15px;
            margin-bottom: 20px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        .stats {
            background-color: #ecf0f1;
            padding: 10px;
            border-radius: 3px;
            margin: 10px 0;
        }
        .error {
            color: #e74c3c;
            font-weight: bold;
            padding: 10px;
            background-color: #fadbd8;
            border-radius: 3px;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 10px;
        }
        th {
            background-color: #34495e;
            color: white;
            padding: 10px;
            text-align: left;
            font-weight: bold;
        }
        td {
            padding: 8px;
            border-bottom: 1px solid #ddd;
        }
        tr:hover {
            background-color: #f9f9f9;
        }
        .cli { color: #27ae60; font-weight: bold; }
        .netconf { color: #2980b9; font-weight: bold; }
        .other { color: #95a5a6; }
    </style>
</head>
<body>
""")
        f.write("<h1>CLI Usage Tracker - Detailed Report</h1>\n")
        f.write(f"<p class='meta'>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>\n")

        for result in sorted(results, key=lambda r: r["hostname"]):
            hostname = result["hostname"]
            vendor = result["vendor"]
            error = result["error"]
            commits = result["commits"]

            f.write("<div class='device-section'>\n")
            f.write(f"<h2>{hostname} <span style='color: #7f8c8d; font-size: 0.8em;'>(Vendor: {vendor})</span></h2>\n")

            if error:
                f.write(f"<p class='error'>ERROR: {error}</p>\n")
                f.write("</div>\n")
                continue

            if not commits:
                f.write("<p>No commits found</p>\n")
                f.write("</div>\n")
                continue

            # Calculate device-specific stats
            cli_count = sum(1 for c in commits if "cli" in c["via"])
            netconf_count = sum(1 for c in commits if "netconf" in c["via"] or "grpc" in c["via"])
            other_count = len(commits) - cli_count - netconf_count

            f.write("<div class='stats'>\n")
            f.write(f"<strong>Total Commits:</strong> {len(commits)}<br>\n")
            f.write(f"<span class='cli'>CLI: {cli_count} ({cli_count/len(commits)*100:.1f}%)</span> | ")
            f.write(f"<span class='netconf'>NETCONF: {netconf_count} ({netconf_count/len(commits)*100:.1f}%)</span>")
            if other_count > 0:
                f.write(f" | <span class='other'>Other: {other_count} ({other_count/len(commits)*100:.1f}%)</span>")
            f.write("\n</div>\n")

            f.write("<table>\n")
            f.write("<tr><th>Timestamp</th><th>User</th><th>Method</th><th>Comment</th></tr>\n")

            for commit in commits:
                method_class = "cli" if "cli" in commit["via"] else ("netconf" if "netconf" in commit["via"] or "grpc" in commit["via"] else "other")
                f.write(
                    f"<tr><td>{commit['datetime']}</td><td>{commit['user']}</td>"
                    f"<td class='{method_class}'>{commit['via']}</td>"
                    f"<td>{commit['comment'][:80]}</td></tr>\n"
                )

            f.write("</table>\n")
            f.write("</div>\n")

        f.write("</body>\n</html>\n")


def write_summary_report_html(stats: dict, filepath: Path):
    """Write fleet-wide summary statistics in HTML format."""
    with open(filepath, "w") as f:
        f.write("""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>CLI Usage Tracker - Summary Report</title>
    <style>
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }
        h1 {
            color: #2c3e50;
            border-bottom: 3px solid #3498db;
            padding-bottom: 10px;
        }
        h2 {
            color: #34495e;
            margin-top: 30px;
        }
        .meta {
            color: #7f8c8d;
            font-style: italic;
            margin-bottom: 20px;
        }
        .summary-box {
            background-color: white;
            border: 1px solid #ddd;
            border-radius: 5px;
            padding: 20px;
            margin-bottom: 20px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        .stat-row {
            padding: 8px 0;
            border-bottom: 1px solid #ecf0f1;
        }
        .stat-label {
            font-weight: bold;
            display: inline-block;
            width: 200px;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 10px;
            background-color: white;
        }
        th {
            background-color: #34495e;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: bold;
        }
        td {
            padding: 10px 12px;
            border-bottom: 1px solid #ddd;
        }
        tr:hover {
            background-color: #f9f9f9;
        }
        .cli { color: #27ae60; font-weight: bold; }
        .netconf { color: #2980b9; font-weight: bold; }
    </style>
</head>
<body>
""")
        f.write("<h1>CLI Usage Tracker - Summary Report</h1>\n")
        f.write(f"<p class='meta'>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>\n")

        # Overall statistics
        f.write("<div class='summary-box'>\n")
        f.write("<h2>Fleet-Wide Statistics</h2>\n")
        f.write(f"<div class='stat-row'><span class='stat-label'>Total Devices:</span> {stats['total_devices']}</div>\n")
        f.write(f"<div class='stat-row'><span class='stat-label'>Devices Queried:</span> {stats['devices_queried']}</div>\n")
        f.write(f"<div class='stat-row'><span class='stat-label'>Devices Failed:</span> {stats['devices_failed']}</div>\n")
        f.write(f"<div class='stat-row'><span class='stat-label'>Total Commits:</span> {stats['total_commits']}</div>\n")

        if stats["total_commits"] > 0:
            cli_pct = stats["cli_commits"] / stats["total_commits"] * 100
            netconf_pct = stats["netconf_commits"] / stats["total_commits"] * 100
            other_pct = stats["other_commits"] / stats["total_commits"] * 100

            f.write(f"<div class='stat-row'><span class='stat-label cli'>CLI Commits:</span> {stats['cli_commits']} ({cli_pct:.1f}%)</div>\n")
            f.write(f"<div class='stat-row'><span class='stat-label netconf'>NETCONF Commits:</span> {stats['netconf_commits']} ({netconf_pct:.1f}%)</div>\n")
            if stats["other_commits"] > 0:
                f.write(f"<div class='stat-row'><span class='stat-label'>Other Commits:</span> {stats['other_commits']} ({other_pct:.1f}%)</div>\n")

        f.write("</div>\n")

        # By vendor
        f.write("<div class='summary-box'>\n")
        f.write("<h2>Breakdown by Vendor</h2>\n")
        f.write("<table>\n")
        f.write("<tr><th>Vendor</th><th>Total</th><th>CLI</th><th>CLI %</th><th>NETCONF</th><th>NETCONF %</th><th>Other</th></tr>\n")

        for vendor in sorted(stats["by_vendor"].keys()):
            v_stats = stats["by_vendor"][vendor]
            total = v_stats["total"]
            if total > 0:
                cli_pct = v_stats["cli"] / total * 100
                netconf_pct = v_stats["netconf"] / total * 100
                f.write(
                    f"<tr><td><strong>{vendor}</strong></td><td>{total}</td>"
                    f"<td class='cli'>{v_stats['cli']}</td><td>{cli_pct:.1f}%</td>"
                    f"<td class='netconf'>{v_stats['netconf']}</td><td>{netconf_pct:.1f}%</td>"
                    f"<td>{v_stats['other']}</td></tr>\n"
                )

        f.write("</table>\n")
        f.write("</div>\n")

        # By user
        f.write("<div class='summary-box'>\n")
        f.write("<h2>Breakdown by User (Top 20)</h2>\n")
        f.write("<table>\n")
        f.write("<tr><th>User</th><th>Total</th><th>CLI</th><th>CLI %</th><th>NETCONF</th><th>NETCONF %</th><th>Other</th></tr>\n")

        # Sort by total commits descending
        sorted_users = sorted(
            stats["by_user"].items(),
            key=lambda x: x[1]["total"],
            reverse=True
        )[:20]

        for user, u_stats in sorted_users:
            total = u_stats["total"]
            if total > 0:
                cli_pct = u_stats["cli"] / total * 100
                netconf_pct = u_stats["netconf"] / total * 100
                f.write(
                    f"<tr><td><strong>{user}</strong></td><td>{total}</td>"
                    f"<td class='cli'>{u_stats['cli']}</td><td>{cli_pct:.1f}%</td>"
                    f"<td class='netconf'>{u_stats['netconf']}</td><td>{netconf_pct:.1f}%</td>"
                    f"<td>{u_stats['other']}</td></tr>\n"
                )

        f.write("</table>\n")
        f.write("</div>\n")

        f.write("</body>\n</html>\n")


def write_cli_only_report_html(results: list[dict], filepath: Path):
    """Write HTML report showing ONLY CLI commits (filters out NETCONF-only entries)."""
    with open(filepath, "w") as f:
        f.write("""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>CLI Usage Tracker - CLI-Only Report</title>
    <style>
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }
        h1 {
            color: #2c3e50;
            border-bottom: 3px solid #27ae60;
            padding-bottom: 10px;
        }
        h2 {
            color: #34495e;
            background-color: #d5f4e6;
            padding: 10px;
            border-left: 5px solid #27ae60;
            margin-top: 30px;
        }
        .meta {
            color: #7f8c8d;
            font-style: italic;
            margin-bottom: 20px;
        }
        .note {
            background-color: #fff3cd;
            border: 1px solid #ffc107;
            padding: 10px;
            border-radius: 5px;
            margin-bottom: 20px;
        }
        .device-section {
            background-color: white;
            border: 1px solid #ddd;
            border-radius: 5px;
            padding: 15px;
            margin-bottom: 20px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        .stats {
            background-color: #d5f4e6;
            padding: 10px;
            border-radius: 3px;
            margin: 10px 0;
            color: #27ae60;
            font-weight: bold;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 10px;
        }
        th {
            background-color: #27ae60;
            color: white;
            padding: 10px;
            text-align: left;
            font-weight: bold;
        }
        td {
            padding: 8px;
            border-bottom: 1px solid #ddd;
        }
        tr:hover {
            background-color: #f9f9f9;
        }
        .cli { color: #27ae60; font-weight: bold; }
        .summary {
            background-color: white;
            border: 2px solid #27ae60;
            border-radius: 5px;
            padding: 20px;
            margin-top: 30px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
    </style>
</head>
<body>
""")
        f.write("<h1>CLI Usage Tracker - CLI-Only Report</h1>\n")
        f.write(f"<p class='meta'>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>\n")
        f.write("<div class='note'><strong>NOTE:</strong> This report shows ONLY commits made via CLI (NETCONF commits filtered out)</div>\n")

        cli_device_count = 0
        total_cli_commits = 0

        for result in sorted(results, key=lambda r: r["hostname"]):
            hostname = result["hostname"]
            vendor = result["vendor"]
            error = result["error"]
            commits = result["commits"]

            if error or not commits:
                continue

            # Filter to only CLI commits
            cli_commits = [c for c in commits if "cli" in c["via"]]

            if not cli_commits:
                continue

            cli_device_count += 1
            total_cli_commits += len(cli_commits)

            f.write("<div class='device-section'>\n")
            f.write(f"<h2>{hostname} <span style='color: #7f8c8d; font-size: 0.8em;'>(Vendor: {vendor})</span></h2>\n")
            f.write(f"<div class='stats'>CLI Commits: {len(cli_commits)}</div>\n")

            f.write("<table>\n")
            f.write("<tr><th>Timestamp</th><th>User</th><th>Method</th><th>Comment</th></tr>\n")

            for commit in cli_commits:
                f.write(
                    f"<tr><td>{commit['datetime']}</td><td>{commit['user']}</td>"
                    f"<td class='cli'>{commit['via']}</td>"
                    f"<td>{commit['comment'][:80]}</td></tr>\n"
                )

            f.write("</table>\n")
            f.write("</div>\n")

        f.write("<div class='summary'>\n")
        f.write("<h2>CLI-Only Summary</h2>\n")
        f.write(f"<div class='stat-row'><span class='stat-label'>Devices with CLI commits:</span> {cli_device_count}</div>\n")
        f.write(f"<div class='stat-row'><span class='stat-label'>Total CLI commits:</span> {total_cli_commits}</div>\n")
        f.write("</div>\n")

        f.write("</body>\n</html>\n")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    global USERNAME, PASSWORD

    parser = argparse.ArgumentParser(
        description="Track CLI vs NETCONF commit usage across network devices",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )
    parser.add_argument(
        "--max-commits",
        type=int,
        default=30,
        help="Maximum number of commits to retrieve per device (default: 30)"
    )
    parser.add_argument(
        "--devices",
        help="Comma-separated list of specific devices to query (default: all from inventory)"
    )
    parser.add_argument(
        "--vendor",
        choices=["Juniper", "Nokia", "Arrcus"],
        help="Query only devices from specific vendor"
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path.cwd(),
        help="Output directory for reports (default: current directory)"
    )
    parser.add_argument(
        "--max-workers",
        type=int,
        default=20,
        help="Maximum concurrent device queries (default: 20)"
    )
    parser.add_argument(
        "--inventory",
        type=Path,
        help="Path to inventory file (CSV from build_inventory.py or Excel). "
             f"Auto-detects: CSV={INVENTORY_CSV}, Excel={INVENTORY_FILE}"
    )
    parser.add_argument(
        "--from-isis",
        action="store_true",
        help="Pull device list dynamically from ISIS database (no inventory file needed)"
    )
    parser.add_argument(
        "--isis-router",
        default=ISIS_ROUTER,
        help=f"Juniper router to query ISIS database from (default: {ISIS_ROUTER})"
    )
    parser.add_argument(
        "--env-file",
        type=Path,
        help="Path to .env file for credentials (default: ~/.env if exists)"
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Enable debug output for troubleshooting connection issues"
    )
    parser.add_argument(
        "--test-device",
        help="Test a single device and show raw output (for debugging). "
             "Format: hostname or hostname:vendor (e.g., use3.am7 or use3.am7:Nokia)"
    )

    args = parser.parse_args()

    # Handle test mode
    if args.test_device:
        parts = args.test_device.split(":")
        test_hostname = parts[0]
        test_vendor = parts[1] if len(parts) > 1 else None

        # Get credentials
        USERNAME = os.environ.get("DEVICE_USERNAME")
        PASSWORD = os.environ.get("DEVICE_PASSWORD")

        if not USERNAME:
            default_user = os.environ.get("USER", "admin")
            USERNAME = input(f"Device username [{default_user}]: ").strip() or default_user
        if not PASSWORD:
            PASSWORD = getpass.getpass("Device password: ")

        print(f"\n{'='*80}")
        print(f"TESTING DEVICE: {test_hostname}")
        if test_vendor:
            print(f"VENDOR: {test_vendor}")
        else:
            test_vendor = guess_vendor_from_hostname(test_hostname)
            print(f"VENDOR (guessed): {test_vendor}")
        print(f"{'='*80}\n")

        # Test NETCONF
        print("Attempting NETCONF connection...")
        try:
            if test_vendor == "Juniper":
                commits = query_juniper_netconf(test_hostname, 30)
            elif test_vendor == "Nokia":
                commits = query_nokia_netconf(test_hostname, 30)
            elif test_vendor == "Arrcus":
                commits = query_arrcus_netconf(test_hostname, 30)
            else:
                print(f"ERROR: Unknown vendor {test_vendor}")
                sys.exit(1)

            print(f"✓ NETCONF SUCCESS: Found {len(commits)} commits")
            if commits:
                print("\nFirst 3 commits:")
                for i, c in enumerate(commits[:3], 1):
                    print(f"  {i}. {c['datetime']} | {c['user']} | {c['via']}")
        except Exception as e:
            print(f"✗ NETCONF FAILED: {e}")
            print("\nAttempting SSH fallback...")

            try:
                if test_vendor == "Juniper":
                    commits = query_juniper_ssh(test_hostname, 30)
                elif test_vendor == "Nokia":
                    commits = query_nokia_ssh(test_hostname, 30)
                elif test_vendor == "Arrcus":
                    commits = query_arrcus_ssh(test_hostname, 30)

                print(f"✓ SSH SUCCESS: Found {len(commits)} commits")
                if commits:
                    print("\nFirst 3 commits:")
                    for i, c in enumerate(commits[:3], 1):
                        print(f"  {i}. {c['datetime']} | {c['user']} | {c['via']}")
                else:
                    print("\nWARNING: SSH succeeded but returned no commits.")
                    print("This could mean:")
                    print("  - Device has no commit history")
                    print("  - Commit history is not enabled")
                    print("  - Command output format is different than expected")
            except Exception as e2:
                print(f"✗ SSH FAILED: {e2}")
                import traceback
                print("\nFull traceback:")
                traceback.print_exc()

        print(f"\n{'='*80}")
        sys.exit(0)

    # Load credentials from .env file if specified or default
    if args.env_file:
        if args.env_file.exists():
            print(f"Loading credentials from {args.env_file}")
            load_dotenv(args.env_file, override=True)
        else:
            print(f"WARNING: Specified .env file not found: {args.env_file}", file=sys.stderr)
    else:
        # Try default .env locations
        default_env_paths = [
            Path.cwd() / ".env",
            Path.home() / ".env",
        ]
        for env_path in default_env_paths:
            if env_path.exists():
                print(f"Loading credentials from {env_path}")
                load_dotenv(env_path, override=True)
                break

    # Get credentials from environment or prompt
    USERNAME = os.environ.get("DEVICE_USERNAME")
    PASSWORD = os.environ.get("DEVICE_PASSWORD")

    if not USERNAME:
        default_user = os.environ.get("USER", "admin")
            USERNAME = input(f"Device username [{default_user}]: ").strip() or default_user

    if not PASSWORD:
        PASSWORD = getpass.getpass("Device password: ")

    if not PASSWORD:
        print("ERROR: Password is required", file=sys.stderr)
        sys.exit(1)

    print(f"Authenticated as: {USERNAME}")

    # Determine device source: ISIS discovery or inventory file
    if args.from_isis:
        # Pull device list from ISIS dynamically
        hostnames = pull_isis_hostnames(router=args.isis_router, port=ISIS_PORT)
        print(f"\nGuessing vendors from hostname patterns...")
        inventory = {}
        for hostname in hostnames:
            vendor = guess_vendor_from_hostname(hostname)
            inventory[hostname] = vendor

        # Report vendor distribution
        from collections import Counter
        vendor_counts = Counter(inventory.values())
        print(f"\nVendor distribution (guessed from hostnames):")
        for vendor in sorted(vendor_counts.keys()):
            print(f"  {vendor}: {vendor_counts[vendor]}")

    else:
        # Auto-detect inventory file if not specified
        if args.inventory is None:
            if Path(INVENTORY_CSV).exists():
                args.inventory = Path(INVENTORY_CSV)
                print(f"Auto-detected CSV inventory: {args.inventory}")
            elif Path(INVENTORY_FILE).exists():
                args.inventory = Path(INVENTORY_FILE)
                print(f"Auto-detected Excel inventory: {args.inventory}")
            else:
                print(f"ERROR: No inventory file found. Checked:", file=sys.stderr)
                print(f"  - {INVENTORY_CSV}", file=sys.stderr)
                print(f"  - {INVENTORY_FILE}", file=sys.stderr)
                print(f"\nOptions:", file=sys.stderr)
                print(f"  1. Generate inventory: cd ~/NEO-ENG-TOOLS/build_inventory && python3 build_inventory.py", file=sys.stderr)
                print(f"  2. Use ISIS discovery: python3 cli_usage_tracker.py --from-isis", file=sys.stderr)
                sys.exit(1)

        # Load inventory
        if not args.inventory.exists():
            print(f"ERROR: Inventory file not found: {args.inventory}", file=sys.stderr)
            sys.exit(1)

        print(f"Loading inventory from {args.inventory}...")
        inventory = load_inventory(str(args.inventory))

    # Filter devices
    if args.devices:
        device_list = [d.strip() for d in args.devices.split(",")]
        devices = {h: inventory.get(h, "unknown") for h in device_list}
    else:
        devices = inventory

    if args.vendor:
        devices = {h: v for h, v in devices.items() if v == args.vendor}

    if not devices:
        print("ERROR: No devices to query", file=sys.stderr)
        sys.exit(1)

    print(f"Querying {len(devices)} devices (max {args.max_commits} commits per device)...\n")

    # Query devices in parallel
    results = []
    with ThreadPoolExecutor(max_workers=args.max_workers) as pool:
        futures = {}
        for hostname, vendor in devices.items():
            futures[pool.submit(query_device, hostname, vendor, args.max_commits, args.debug)] = (hostname, vendor)

        for future in as_completed(futures):
            hostname, vendor = futures[future]
            result = future.result()
            results.append(result)

            status = "OK" if not result["error"] else "FAIL"
            commit_count = len(result["commits"])
            cli_count = sum(1 for c in result["commits"] if "cli" in c["via"])
            netconf_count = sum(1 for c in result["commits"] if "netconf" in c["via"] or "grpc" in c["via"])

            print(
                f"  {status:4s}  {hostname:<30s} [{vendor:<8s}] "
                f"Commits: {commit_count:3d}  CLI: {cli_count:3d}  NETCONF: {netconf_count:3d}"
                f"{('  ERROR: ' + result['error'][:40]) if result['error'] else ''}"
            )

    print(f"\nGenerating reports...")

    # Generate reports
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    args.output_dir.mkdir(parents=True, exist_ok=True)

    # Text reports
    detailed_report = args.output_dir / f"cli_usage_report_{timestamp}.txt"
    summary_report = args.output_dir / f"cli_usage_summary_{timestamp}.txt"
    cli_only_report = args.output_dir / f"cli_usage_cli_only_{timestamp}.txt"
    csv_export = args.output_dir / f"cli_usage_report_{timestamp}.csv"

    # HTML reports
    detailed_report_html = args.output_dir / f"cli_usage_report_{timestamp}.html"
    summary_report_html = args.output_dir / f"cli_usage_summary_{timestamp}.html"
    cli_only_report_html = args.output_dir / f"cli_usage_cli_only_{timestamp}.html"

    write_detailed_report(results, detailed_report)
    stats = analyze_commits(results)
    write_summary_report(stats, summary_report)
    write_cli_only_report(results, cli_only_report)
    write_csv_export(results, csv_export)

    # Generate HTML reports
    write_detailed_report_html(results, detailed_report_html)
    write_summary_report_html(stats, summary_report_html)
    write_cli_only_report_html(results, cli_only_report_html)

    print(f"\nReports generated:")
    print(f"\nText Reports:")
    print(f"  Detailed:    {detailed_report}")
    print(f"  Summary:     {summary_report}")
    print(f"  CLI-Only:    {cli_only_report}")
    print(f"  CSV Export:  {csv_export}")
    print(f"\nHTML Reports:")
    print(f"  Detailed:    {detailed_report_html}")
    print(f"  Summary:     {summary_report_html}")
    print(f"  CLI-Only:    {cli_only_report_html}")
    print(f"\nFleet Summary:")
    print(f"  Total Commits:   {stats['total_commits']}")
    if stats['total_commits'] > 0:
        cli_pct = stats['cli_commits'] / stats['total_commits'] * 100
        netconf_pct = stats['netconf_commits'] / stats['total_commits'] * 100
        print(f"  CLI:             {stats['cli_commits']} ({cli_pct:.1f}%)")
        print(f"  NETCONF:         {stats['netconf_commits']} ({netconf_pct:.1f}%)")


if __name__ == "__main__":
    main()
